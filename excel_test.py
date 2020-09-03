import json

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class Dto:
    pass


def test():
    wb = load_workbook(filename="C:/Users/40668/Desktop/python_test_excel.xlsx")  # 读取excel文件
    sheet_names = wb.sheetnames  # 获取所有的sheet页名字
    sheet_name0: str = sheet_names[0]  # 获取第一个sheet页

    sheet: Worksheet = wb[sheet_name0]  # 按名字获取sheet页
    result: list[dict[str, any]] = to_dict_list(sheet)
    print(result)
    print(type(result))
    j: str = json.dumps(result, ensure_ascii=False)
    print(j)
    print(type(j))


def foreach_row(sheet: Worksheet):
    """
    遍历每行
    :param sheet:
    :return:
    """
    for row in sheet:
        print(row)


def foreach_row_values(sheet: Worksheet):
    """
    遍历每行的数据
    :param sheet:
    :return:
    """
    for row in sheet.values:
        for v in row:
            print(v)


def iter_rows_in_range(sheet: Worksheet, start_row=None, end_row=None, max_col=None):
    """
    按范围遍历
    :param sheet:
    :param start_row: 开始的行数，从1开始
    :param end_row:
    :param max_col: 结束的列数，从1开始
    :return:
    """
    for row in sheet.iter_cols(min_row=start_row, max_row=end_row, max_col=max_col):
        print(row)


def iter_rows_value_in_range(sheet: Worksheet, start_row=None, end_row=None, max_col=None):
    """
    按范围遍历，只包含数据
    :param sheet:
    :param start_row: 开始的行数，从1开始
    :param end_row:
    :param max_col: 结束的列数，从1开始
    :return:
    """
    for row in sheet.iter_cols(min_row=start_row, max_row=end_row, max_col=max_col, values_only=True):
        for v in row:
            print(v)


def to_dto_list(sheet: Worksheet):
    result = []
    title_list = get_title(sheet)
    # title_map = get_title_map(sheet)
    # # 遍历时按照列名取值
    # get_value = lambda _row, name: _row[title_map[name]].value
    for row in sheet.iter_rows(min_row=2):
        dto = Dto()
        for i in range(len(row)):
            value = row[i].value
            att_name = title_list[i]
            dto.__setattr__(att_name, value)
        result.append(dto)
    return result


def to_dict_list(sheet: Worksheet):
    result = []
    title_list = get_title(sheet)
    # title_map = get_title_map(sheet)
    # # 遍历时按照列名取值
    # get_value = lambda _row, name: _row[title_map[name]].value
    for row in sheet.iter_rows(min_row=2):
        d = {}
        for i in range(len(row)):
            value = row[i].value
            att_name = title_list[i]
            d[att_name] = value
        result.append(d)
    return result


def get_title_map(sheet: Worksheet):
    title_rows = sheet[1]
    title_map = {

    }
    for i in range(0, len(title_rows)):
        title_map[title_rows[i].value] = i
    return title_map


def get_title(sheet: Worksheet):
    title_rows = sheet[1]
    title_result = [None]*len(title_rows)
    for i in range(0, len(title_rows)):
        title_result[i] = title_rows[i].value
    return title_result


test()
